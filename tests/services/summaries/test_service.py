import pytest
import pytest_asyncio
import pandas as pd
from api.services.travel.service import SummaryService
from io import BytesIO
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.cell.cell import MergedCell


@pytest.fixture
def sample_bed_night_report_all_columns_dataframe():
    """
    Fixture to create a sample DataFrame including individual bed night entries for testing.
    """
    data = {
        "Primary Traveler": [
            "Test1/Test1",
            "Test2/Test2",
            "Test3/Test3",
            "Test4/Test4",
            "Test5/Test5",
        ],
        "Date In": [
            "2024-05-24",
            "2024-06-24",
            "2024-07-24",
            "2024-08-24",
            "2024-09-24",
        ],
        "Date Out": [
            "2024-05-27",
            "2024-06-27",
            "2024-07-27",
            "2024-08-27",
            "2024-09-27",
        ],
        "# Pax": [2, 2, 4, 8, 8],
        "Bed Nights": [6, 6, 12, 24, 24],
        "Property": [
            "Lodge1",
            "Lodge2",
            "Lodge3",
            "Lodge4",
            "Lodge5",
        ],
        "Portfolio": [
            "Portfolio1",
            "Portfolio2",
            "Portfolio3",
            "Portfolio4",
            "Portfolio5",
        ],
        "Core Destination Name": [
            "Africa",
            "Africa",
            "Africa",
            "Africa",
            "Africa",
        ],
        "Country": [
            "Botswana",
            "Botswana",
            "Botswana",
            "Botswana",
            "Botswana",
        ],
        "Agency": [
            "Agency1",
            "Agency2",
            "Agency3",
            "Agency4",
            "Agency5",
        ],
        "Booking Channel": [
            "BC1",
            "BC2",
            "BC3",
            "BC4",
            "BC5",
        ],
        "Consultant": [
            "Consultant1",
            "Consultant2",
            "Consultant3",
            "Consultant4",
            "Consultant5",
        ],
    }
    df = pd.DataFrame(data)
    return df


@pytest.fixture
def sample_custom_report_property_yearly_dataframe():
    """
    Fixture to create a sample DataFrame including aggregated report results for testing.
    Property granularity: `property_name`
    Time granularity: `year`
    """
    data = {
        "Property": ["Hotel A", "Hotel B", "Hotel C", "Hotel D", "Hotel E"],
        "2024": [0, 2, 3, 8, 2],
        "2025": [1, 3, 4, 5, 6],
        "2026": [2, 4, 5, 6, 7],
    }
    df = pd.DataFrame(data)
    return df


@pytest_asyncio.fixture
async def summary_service():
    """
    Fixture to instantiate the ReportGenerator class.
    """
    return SummaryService()


@pytest.mark.asyncio
async def sample_custom_excel_property_yearly(
    sample_custom_report_property_yearly_dataframe, summary_service
):
    # Call the write_excel function with include_total_column=True
    excel_stream = await summary_service.write_excel(
        df=sample_custom_report_property_yearly_dataframe,
        report_title="Aggregated Report",
        include_total_column=True,
    )

    # Load the workbook from the BytesIO stream
    wb = load_workbook(filename=BytesIO(excel_stream.getvalue()))
    sheet = wb.active
    assert sheet is not None, "No sheet resulted from the write_excel function"

    # Define expected row numbers
    header_row = 3
    data_start_row = 4
    data_end_row = (
        data_start_row + len(sample_custom_report_property_yearly_dataframe) - 1
    )
    total_row = data_end_row + 1

    # ----------------------------
    # 1. Validate Headers
    # ----------------------------

    # Column A: PROPERTY

    header_property = sheet.cell(row=header_row, column=1)
    assert (
        header_property.value == "PROPERTY"
    ), "Header in column A should be 'PROPERTY'"
    assert (
        header_property.number_format == "General"
    ), "PROPERTY header should have 'General' number format"
    assert (
        header_property.alignment.horizontal == "left"
    ), "PROPERTY header should be left-aligned"

    # Column B, C, D: 2024, 2025, 2026 (Numeric Header)
    for i in range(2, 4):
        header_cell = sheet.cell(row=header_row, column=i)
        assert (
            header_cell.value == 0
        ), f"Header {i} should be stored as a number, not text."
        assert isinstance(header_cell.value, int), f"Header {i} should be an integer."
        assert (
            header_cell.number_format == "0"
        ), f"Header {i} should have '0' number format"
        assert (
            header_cell.alignment.horizontal == "center"
        ), f"Header {i} should be center-aligned"

    # Column E: TOTAL
    header_total = sheet.cell(row=header_row, column=5)
    assert header_total.value == "TOTAL", "Header in column E should be 'TOTAL'"
    assert (
        header_total.number_format == "General"
    ), "TOTAL header should have 'General' number format"
    assert (
        header_total.alignment.horizontal == "center"
    ), "TOTAL header should be center-aligned"


@pytest.mark.asyncio
async def test_write_excel_bed_night_all_columns(
    sample_bed_night_report_all_columns_dataframe, summary_service
):
    # Call the write_excel function with include_total_column=False
    excel_stream = await summary_service.write_excel(
        df=sample_bed_night_report_all_columns_dataframe,
        report_title="Bed Night Report",
        include_total_column=False,
    )

    # Load the workbook from the BytesIO stream
    wb = load_workbook(filename=BytesIO(excel_stream.getvalue()))
    sheet = wb.active
    assert sheet is not None, "No sheet resulted from the write_excel function"

    # Define expected row numbers
    header_row = 3
    data_start_row = 4
    data_end_row = (
        data_start_row + len(sample_bed_night_report_all_columns_dataframe) - 1
    )
    total_row = data_end_row + 1

    # Calculate expected number of columns (without TOTAL)
    num_columns = len(sample_bed_night_report_all_columns_dataframe.columns)
    assert (
        sheet.max_column == num_columns
    ), f"Expected {num_columns} columns, found {sheet.max_column}"

    # ----------------------------
    # 1. Validate Headers
    # ----------------------------

    # Column A: PRIMARY TRAVELER
    header_property_traveler = sheet.cell(row=header_row, column=1)
    assert (
        header_property_traveler.value == "PRIMARY TRAVELER"
    ), "Header in column A should be 'PRIMARY TRAVELER'"
    assert (
        header_property_traveler.number_format == "General"
    ), "PRIMARY TRAVELER header should have 'General' number format"
    assert (
        header_property_traveler.alignment.horizontal == "center"
    ), "PRIMARY TRAVELER header should be center-aligned"

    # Column B: DATE IN
    header_date_in = sheet.cell(row=header_row, column=2)
    assert header_date_in.value == "DATE IN", "Header in column B should be 'DATE IN'"
    assert (
        header_date_in.number_format == "General"
    ), "DATE IN header should have 'General' number format"
    assert (
        header_date_in.alignment.horizontal == "center"
    ), "DATE IN header should be center-aligned"

    # Column C: DATE OUT
    header_date_out = sheet.cell(row=header_row, column=3)
    assert (
        header_date_out.value == "DATE OUT"
    ), "Header in column C should be 'DATE OUT'"
    assert (
        header_date_out.number_format == "General"
    ), "DATE OUT header should have 'General' number format"
    assert (
        header_date_out.alignment.horizontal == "center"
    ), "DATE OUT header should be center-aligned"

    # Column D: # PAX
    header_pax = sheet.cell(row=header_row, column=4)
    assert header_pax.value == "# PAX", "Header in column D should be '# PAX'"
    assert (
        header_pax.number_format == "General"
    ), "# PAX header should have 'General' number format"
    assert (
        header_pax.alignment.horizontal == "center"
    ), "# PAX header should be center-aligned"

    # Column E: BED NIGHTS
    header_bed_nights = sheet.cell(row=header_row, column=5)
    assert (
        header_bed_nights.value == "BED NIGHTS"
    ), "Header in column E should be 'BED NIGHTS'"
    assert (
        header_bed_nights.number_format == "General"
    ), "BED NIGHTS header should have 'General' number format"
    assert (
        header_bed_nights.alignment.horizontal == "center"
    ), "BED NIGHTS header should be center-aligned"

    # Column F: PROPERTY
    header_property = sheet.cell(row=header_row, column=6)
    assert (
        header_property.value == "PROPERTY"
    ), "Header in column F should be 'PROPERTY'"
    assert (
        header_property.number_format == "General"
    ), "PROPERTY header should have 'General' number format"
    assert (
        header_property.alignment.horizontal == "center"
    ), "PROPERTY header should be center-aligned"

    # Column G: PORTFOLIO
    header_portfolio = sheet.cell(row=header_row, column=7)
    assert (
        header_portfolio.value == "PORTFOLIO"
    ), "Header in column G should be 'PORTFOLIO'"
    assert (
        header_portfolio.number_format == "General"
    ), "PORTFOLIO header should have 'General' number format"
    assert (
        header_portfolio.alignment.horizontal == "center"
    ), "PORTFOLIO header should be center-aligned"

    # Column H: CORE DESTINATION NAME
    header_core_destination = sheet.cell(row=header_row, column=8)
    assert (
        header_core_destination.value == "CORE DESTINATION NAME"
    ), "Header in column H should be 'CORE DESTINATION NAME'"
    assert (
        header_core_destination.number_format == "General"
    ), "CORE DESTINATION NAME header should have 'General' number format"
    assert (
        header_core_destination.alignment.horizontal == "center"
    ), "CORE DESTINATION NAME header should be center-aligned"

    # Column I: COUNTRY
    header_country = sheet.cell(row=header_row, column=9)
    assert header_country.value == "COUNTRY", "Header in column I should be 'COUNTRY'"
    assert (
        header_country.number_format == "General"
    ), "COUNTRY header should have 'General' number format"
    assert (
        header_country.alignment.horizontal == "center"
    ), "COUNTRY header should be center-aligned"

    # Column J: AGENCY
    header_agency = sheet.cell(row=header_row, column=10)
    assert header_agency.value == "AGENCY", "Header in column J should be 'AGENCY'"
    assert (
        header_agency.number_format == "General"
    ), "AGENCY header should have 'General' number format"
    assert (
        header_agency.alignment.horizontal == "center"
    ), "AGENCY header should be center-aligned"

    # Column K: BOOKING CHANNEL
    header_booking_channel = sheet.cell(row=header_row, column=11)
    assert (
        header_booking_channel.value == "BOOKING CHANNEL"
    ), "Header in column K should be 'BOOKING CHANNEL'"
    assert (
        header_booking_channel.number_format == "General"
    ), "BOOKING CHANNEL header should have 'General' number format"
    assert (
        header_booking_channel.alignment.horizontal == "center"
    ), "BOOKING CHANNEL header should be center-aligned"

    # Column L: CONSULTANT
    header_consultant = sheet.cell(row=header_row, column=12)
    assert (
        header_consultant.value == "CONSULTANT"
    ), "Header in column L should be 'CONSULTANT'"
    assert (
        header_consultant.number_format == "General"
    ), "CONSULTANT header should have 'General' number format"
    assert (
        header_consultant.alignment.horizontal == "center"
    ), "CONSULTANT header should be center-aligned"

    # ----------------------------
    # 2. Validate Data Cells
    # ----------------------------

    for index, row_data in sample_bed_night_report_all_columns_dataframe.iterrows():
        excel_row = data_start_row + index

        # Column A: PRIMARY TRAVELER
        cell_primary_traveler = sheet.cell(row=excel_row, column=1)
        expected_primary_traveler = row_data["Primary Traveler"]
        assert (
            cell_primary_traveler.value == expected_primary_traveler
        ), f"Cell A{excel_row} should be '{expected_primary_traveler}'"

        # Column B: DATE IN
        cell_date_in = sheet.cell(row=excel_row, column=2)
        expected_date_in = row_data["Date In"]
        assert (
            cell_date_in.value == expected_date_in
        ), f"Cell B{excel_row} should be '{expected_date_in}'"

        # Column C: DATE OUT
        cell_date_out = sheet.cell(row=excel_row, column=3)
        expected_date_out = row_data["Date Out"]
        assert (
            cell_date_out.value == expected_date_out
        ), f"Cell C{excel_row} should be '{expected_date_out}'"

        # Column D: # PAX
        cell_pax = sheet.cell(row=excel_row, column=4)
        expected_pax = row_data["# Pax"]
        assert (
            cell_pax.value == expected_pax
        ), f"Cell D{excel_row} should be {expected_pax}"
        assert isinstance(
            cell_pax.value, int
        ), f"Cell D{excel_row} should be an integer."
        assert (
            cell_pax.number_format == "General"
        ), f"Cell D{excel_row} should have 'General' number format"
        assert (
            cell_pax.alignment.horizontal == "center"
        ), f"Cell D{excel_row} should be center-aligned"

        # Column E: BED NIGHTS
        cell_bed_nights = sheet.cell(row=excel_row, column=5)
        expected_bed_nights = row_data["Bed Nights"]
        assert (
            cell_bed_nights.value == expected_bed_nights
        ), f"Cell E{excel_row} should be {expected_bed_nights}"
        assert isinstance(
            cell_bed_nights.value, int
        ), f"Cell E{excel_row} should be an integer."
        assert (
            cell_bed_nights.number_format == "General"
        ), f"Cell E{excel_row} should have 'General' number format"
        assert (
            cell_bed_nights.alignment.horizontal == "center"
        ), f"Cell E{excel_row} should be center-aligned"

        # Column F: PROPERTY
        cell_property = sheet.cell(row=excel_row, column=6)
        expected_property = row_data["Property"]
        assert (
            cell_property.value == expected_property
        ), f"Cell F{excel_row} should be '{expected_property}'"

        # Column G: PORTFOLIO
        cell_property_portfolio = sheet.cell(row=excel_row, column=7)
        expected_property_portfolio = row_data["Portfolio"]
        assert (
            cell_property_portfolio.value == expected_property_portfolio
        ), f"Cell G{excel_row} should be '{expected_property_portfolio}'"

        # Column H: CORE DESTINATION NAME
        cell_core_destination = sheet.cell(row=excel_row, column=8)
        expected_core_destination = row_data["Core Destination Name"]
        assert (
            cell_core_destination.value == expected_core_destination
        ), f"Cell H{excel_row} should be '{expected_core_destination}'"

        # Column I: COUNTRY
        cell_country = sheet.cell(row=excel_row, column=9)
        expected_country = row_data["Country"]
        assert (
            cell_country.value == expected_country
        ), f"Cell I{excel_row} should be '{expected_country}'"

        # Column J: AGENCY
        cell_agency = sheet.cell(row=excel_row, column=10)
        expected_agency = row_data["Agency"]
        assert (
            cell_agency.value == expected_agency
        ), f"Cell J{excel_row} should be '{expected_agency}'"

        # Column K: BOOKING CHANNEL
        cell_booking_channel = sheet.cell(row=excel_row, column=11)
        expected_booking_channel = row_data["Booking Channel"]
        assert (
            cell_booking_channel.value == expected_booking_channel
        ), f"Cell K{excel_row} should be '{expected_booking_channel}'"

        # Column L: CONSULTANT
        cell_consultant = sheet.cell(row=excel_row, column=12)
        expected_consultant = row_data["Consultant"]
        assert (
            cell_consultant.value == expected_consultant
        ), f"Cell L{excel_row} should be '{expected_consultant}'"

    # ----------------------------
    # 3. Validate TOTAL Row
    # ----------------------------

    # TOTAL Row should exist at row=total_row
    # Since include_total_column=False, there should be no TOTAL column (max_column=12)

    # Column A: TOTAL label
    cell_total_label = sheet.cell(row=total_row, column=1)
    assert (
        cell_total_label.value == "TOTAL"
    ), f"Cell A{total_row} should contain 'TOTAL'"
    assert (
        cell_total_label.alignment.horizontal == "right"
    ), f"Cell A{total_row} should be right-aligned"

    # Columns B to L: Sum of numeric columns where applicable
    # In this DataFrame, numeric columns are: "# Pax", "Bed Nights"

    # Mapping column indices to column names
    column_mapping = {
        4: "# PAX",
        5: "BED NIGHTS",
    }

    for col_num, _ in column_mapping.items():
        cell_total = sheet.cell(row=total_row, column=col_num)
        expected_sum_formula = f"=SUM(${get_column_letter(col_num)}${data_start_row}:${get_column_letter(col_num)}${data_end_row})"
        assert (
            cell_total.value == expected_sum_formula
        ), f"Cell {get_column_letter(col_num)}{total_row} should have formula '{expected_sum_formula}'"
        assert (
            cell_total.alignment.horizontal == "center"
        ), f"Cell {get_column_letter(col_num)}{total_row} should be center-aligned"

    # Columns that are non-numeric should be empty in the TOTAL row
    non_numeric_columns = set(range(2, 13)) - set(column_mapping.keys())

    for col_num in non_numeric_columns:
        cell_total = sheet.cell(row=total_row, column=col_num)
        assert (
            cell_total.value == None
        ), f"Cell {get_column_letter(col_num)}{total_row} should be None as it's a non-numeric column"

    # ----------------------------
    # 4. Validate No TOTAL Column Exists
    # ----------------------------

    # The DataFrame has 12 columns, and since include_total_column=False, Excel should also have 12 columns
    assert (
        sheet.max_column == num_columns
    ), f"Expected {num_columns} columns, found {sheet.max_column}"

    # Ensure that there is no 'TOTAL' column at the end
    last_column = sheet.cell(row=header_row, column=num_columns)
    assert (
        last_column.value != "TOTAL"
    ), f"Last column should not be 'TOTAL' when include_total_column=False"

    # ----------------------------
    # 5. Validate Formatting (Optional)
    # ----------------------------

    # Check that all header cells have bold font
    for col_num in range(1, num_columns + 1):
        cell = sheet.cell(row=header_row, column=col_num)
        assert (
            cell.font.bold
        ), f"Header cell {get_column_letter(col_num)}{header_row} should have bold font."

    # Check that TOTAL row cells have bold font where applicable
    cell_total_label = sheet.cell(row=total_row, column=1)
    assert (
        cell_total_label.font.bold
    ), f"TOTAL label cell A{total_row} should have bold font."

    for col_num, column_name in column_mapping.items():
        cell_total = sheet.cell(row=total_row, column=col_num)
        assert (
            cell_total.font.bold
        ), f"TOTAL sum cell {get_column_letter(col_num)}{total_row} should have bold font."

    # Ensure non-sum cells in TOTAL row are not bold
    for col_num in non_numeric_columns:
        cell_total = sheet.cell(row=total_row, column=col_num)
        assert (
            not cell_total.font.bold
        ), f"Non-sum cell {get_column_letter(col_num)}{total_row} should not have bold font."

    # ----------------------------
    # 6. Validate Borders (Optional)
    # ----------------------------

    # All cells have thin borders
    from openpyxl.styles import Border, Side

    thin_side = Side(style="thin", color="000000")
    thin_border = Border(
        left=thin_side, right=thin_side, top=thin_side, bottom=thin_side
    )

    for row in range(header_row, total_row + 1):
        for col in range(1, num_columns + 1):
            cell = sheet.cell(row=row, column=col)
            assert (
                cell.border.left.style == "thin"
                and cell.border.right.style == "thin"
                and cell.border.top.style == "thin"
                and cell.border.bottom.style == "thin"
            ), f"Cell {get_column_letter(col)}{row} should have thin borders on all sides."

    # ----------------------------
    # 7. Validate Column Widths (Optional)
    # ----------------------------

    # Ensure that the widths are set to a reasonable value
    for col in sheet.columns:
        max_length = 0
        # Initialize the column variable
        column = None

        # Find the first non-merged cell in the column to get the column letter
        for cell in col:
            if not isinstance(cell, MergedCell):
                column = cell.column_letter
                break

        if column is None:
            # All cells in the column are merged; skip adjusting this column
            continue

        for cell in col:
            if isinstance(cell, MergedCell):
                # Skip MergedCell objects
                continue
            try:
                cell_value = str(cell.value) if cell.value is not None else ""
                cell_length = len(cell_value)
                if cell_length > max_length:
                    max_length = cell_length
            except:
                pass
        adjusted_width = sheet.column_dimensions[column].width
        assert (
            adjusted_width >= max_length + 2
        ), f"Column {column} width should be at least {max_length + 2}, found {adjusted_width}"

    # ----------------------------
    # 8. Validate Footer
    # ----------------------------

    footer_row = total_row + 1
    footer_cell = sheet.cell(row=footer_row, column=1)
    assert (
        footer_cell.value == "Travel Beyond Confidential"
    ), f"Footer cell A{footer_row} should contain 'Travel Beyond Confidential'"
    assert (
        footer_cell.alignment.horizontal == "center"
    ), f"Footer cell A{footer_row} should be center-aligned"
    assert footer_cell.font.bold, f"Footer cell A{footer_row} should have bold font."

    # ----------------------------
    # 9. Ensure No Formula Errors
    # ----------------------------

    # Ensure that sum formulas do not reference header rows
    for col_num, column_name in column_mapping.items():
        cell_total = sheet.cell(row=total_row, column=col_num)
        # The sum range should start at data_start_row and end at data_end_row
        expected_sum_range = f"${get_column_letter(col_num)}${data_start_row}:${get_column_letter(col_num)}${data_end_row}"
        expected_sum_formula = f"=SUM({expected_sum_range})"
        assert (
            cell_total.value == expected_sum_formula
        ), f"Cell {get_column_letter(col_num)}{total_row} should have formula '{expected_sum_formula}'"

    # Ensure that non-numeric columns in TOTAL row are empty
    for col_num in non_numeric_columns:
        cell_total = sheet.cell(row=total_row, column=col_num)
        assert (
            cell_total.value == None
        ), f"Cell {get_column_letter(col_num)}{total_row} should be None as it's a non-numeric column"

    # ----------------------------
    # 10. Summary Assertions
    # ----------------------------

    # Ensure the TOTAL row has the correct position
    assert (
        sheet.cell(row=total_row, column=1).value == "TOTAL"
    ), f"Row {total_row} should start with 'TOTAL' in column A"

    # Ensure that all sum formulas are correctly set
    for col_num, _ in column_mapping.items():
        cell_total = sheet.cell(row=total_row, column=col_num)
        sum_range = f"${get_column_letter(col_num)}${data_start_row}:${get_column_letter(col_num)}${data_end_row}"
        expected_sum_formula = f"=SUM({sum_range})"
        assert (
            cell_total.value == expected_sum_formula
        ), f"Cell {get_column_letter(col_num)}{total_row} should have formula '{expected_sum_formula}'"
