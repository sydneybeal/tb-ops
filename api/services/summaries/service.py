# Copyright 2024 SH

# Licensed under the Apache License, Version 2.0 (the "License");
# you may not use this file except in compliance with the License.
# You may obtain a copy of the License at

#     http://www.apache.org/licenses/LICENSE-2.0

# Unless required by applicable law or agreed to in writing, software
# distributed under the License is distributed on an "AS IS" BASIS,
# WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
# See the License for the specific language governing permissions and
# limitations under the License.

"""Services for interacting with travel entries."""
from collections import Counter, defaultdict
from datetime import date, datetime
from itertools import groupby
from operator import attrgetter
from typing import Sequence, List, Optional, Set
from uuid import UUID
from io import BytesIO
from xml.etree.ElementInclude import include
import pandas as pd
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.cell.cell import MergedCell

from api.services.summaries.models import (
    AccommodationLogSummary,
    AgencySummary,
    BedNightReport,
    BookingChannelSummary,
    BreakdownItem,
    CountrySummary,
    PortfolioSummary,
    PropertyDetailSummary,
    PropertySummary,
    ReportAggregations,
    ReportInput,
    Overlap,
    TripSummary,
)
from api.services.summaries.repository.postgres import PostgresSummaryRepository

from api.services.travel.models import Trip

FONT_NAME = "Brandon Grotesque"


class SummaryService:
    """Service for interfacing with the travel repository."""

    def __init__(self):
        """Initializes with a configured repository."""
        self._repo = PostgresSummaryRepository()

    # BedNightReport
    async def get_bed_night_report(self, labels: dict) -> BedNightReport:
        """Generates a BedNightReport based on input criteria."""
        accommodation_logs = await self._repo.get_accommodation_logs_by_filter(
            labels, exclude_fam=True
        )
        report = self.generate_report(accommodation_logs, labels)
        return report

    def generate_report(
        self, accommodation_logs: Sequence[AccommodationLogSummary], labels: dict
    ) -> BedNightReport:
        """Performs aggregations on AccommodationLogSummary to generate a report."""
        # Convert labels into ReportInput
        report_input = ReportInput(**labels)

        # Initialize aggregation variables
        total_bed_nights = 0
        country_breakdown_counter = Counter()
        month_breakdown_counter = Counter()
        portfolio_breakdown_counter = Counter()
        property_breakdown_counter = Counter()
        consultant_breakdown_counter = Counter()
        core_destination_breakdown_counter = Counter()

        largest_booking = {
            "bed_nights": 0,
            "num_nights": 0,
            "num_pax": 0,
            "property_name": "",
        }

        # Aggregate data
        for log in accommodation_logs:
            country_name = log.country_name if log.country_name else "Unknown"
            portfolio_name = (
                log.property_portfolio if log.property_portfolio else "Unknown"
            )
            property_name = log.property_name if log.property_name else "Unknown"
            consultant_name = (
                log.consultant_display_name
                if log.consultant_display_name
                else "Unknown"
            )
            core_destination_name = (
                log.core_destination_name if log.core_destination_name else "Unknown"
            )

            total_bed_nights += log.bed_nights
            country_breakdown_counter[country_name] += log.bed_nights
            month_breakdown_counter[log.date_in.strftime("%Y-%m")] += log.bed_nights
            portfolio_breakdown_counter[portfolio_name] += log.bed_nights
            property_breakdown_counter[property_name] += log.bed_nights
            consultant_breakdown_counter[consultant_name] += log.bed_nights
            core_destination_breakdown_counter[core_destination_name] += log.bed_nights

            if log.bed_nights > largest_booking.get("bed_nights", 0):
                largest_booking = {
                    "bed_nights": log.bed_nights,
                    "num_nights": (log.date_out - log.date_in).days,
                    "num_pax": log.num_pax,
                    "property_name": log.property_name,
                }

        # Create BreakdownItems
        def create_breakdown(items: Counter) -> List[BreakdownItem]:
            total = sum(items.values())
            return [
                BreakdownItem(
                    name=name, bed_nights=count, percentage=(count / total * 100)
                )
                for name, count in items.items()
            ]

        # Populate ReportAggregations
        report_aggregations = ReportAggregations(
            total_bed_nights=total_bed_nights,
            by_country=create_breakdown(country_breakdown_counter),
            by_month=create_breakdown(month_breakdown_counter),
            by_portfolio=create_breakdown(portfolio_breakdown_counter),
            by_property=create_breakdown(property_breakdown_counter),
            by_consultant=create_breakdown(consultant_breakdown_counter),
            by_core_destination=create_breakdown(core_destination_breakdown_counter),
            largest_booking=largest_booking,
        )

        # Create and return the BedNightReport
        return BedNightReport(
            report_inputs=report_input, calculations=report_aggregations
        )

    async def generate_excel_file(
        self,
        labels: dict,
        report_title: str,
        exclude_columns: Optional[List[str]] = None,
    ) -> BytesIO:
        """Generates an excel file with accommodation logs for reporting."""
        accommodation_logs = await self._repo.get_accommodation_logs_by_filter(
            labels, exclude_fam=True
        )
        if not accommodation_logs:
            raise ValueError("No data available for the given filters.")

        # Default set of all columns that might be included from AccommodationLogSummary
        default_columns = [
            "primary_traveler",
            "date_in",
            "date_out",
            "num_pax",
            "bed_nights",
            "property_name",
            "property_portfolio",
            "core_destination_name",
            "country_name",
            "agency_name",
            "booking_channel_name",
            "consultant_name",
        ]

        column_name_mapping = {
            "primary_traveler": "Primary Traveler",
            "date_in": "Date In",
            "date_out": "Date Out",
            "num_pax": "# Pax",
            "bed_nights": "Bed Nights",
            "property_name": "Property",
            "property_portfolio": "Portfolio",
            "core_destination_name": "Core Destination Name",
            "country_name": "Country",
            "agency_name": "Agency",
            "booking_channel_name": "Booking Channel",
            "consultant_name": "Consultant",
        }

        # Remove the columns specified in exclude_columns from default_columns
        if exclude_columns:
            columns_to_include = [
                col for col in default_columns if col not in exclude_columns
            ]
        else:
            columns_to_include = default_columns

        # Generate data for DataFrame
        data = [log.dict() for log in accommodation_logs]

        for log in data:
            log["consultant_name"] = (
                f"{log['consultant_last_name']}/{log['consultant_first_name']}"
            )

        df = pd.DataFrame(data, columns=columns_to_include)
        df.rename(columns=column_name_mapping, inplace=True)

        return await self.write_excel(df, report_title)

    async def generate_custom_excel_file(self, query_params: dict, report_title: str):
        """Generates an excel file with custom calculations for reporting."""
        calculation_type = query_params.pop("calculation_type", None)
        property_granularity = query_params.pop("property_granularity", None)
        time_granularity = query_params.pop("time_granularity", None)

        if not (calculation_type and property_granularity and time_granularity):
            raise ValueError("Missing required parameters for report generation.")

        accommodation_logs = await self._repo.get_accommodation_logs_by_filter(
            query_params, exclude_fam=True
        )
        if not accommodation_logs:
            raise ValueError("No data available for the given filters.")

        results = self.aggregate_custom_report(
            accommodation_logs, calculation_type, property_granularity, time_granularity
        )

        df = self.results_to_dataframe(
            results,
            time_granularity,
            property_granularity,
        )

        return await self.write_excel(df, report_title, include_total_column=True)

    def results_to_dataframe(self, results, time_granularity, property_granularity):
        """Create a pandas DataFrame from the results."""
        data = []
        time_index = set()
        property_index = set()

        granularity_label = {
            "property_name": "Property",
            "country_name": "Country",
            "property_portfolio": "Portfolio",
        }

        for (time_key, prop_key), total in results.items():
            time_index.add(time_key)
            property_index.add(prop_key)
            data.append({"Time": time_key, "Property": prop_key, "Total": total})

        df = pd.DataFrame(data)
        # Pivot DataFrame to get time as columns and properties as rows
        pivot_df = df.pivot(index="Property", columns="Time", values="Total")
        pivot_df = pivot_df.fillna(0)  # Fill missing values with zeros

        # Sort by property names or according to property_granularity
        pivot_df = pivot_df.sort_index()

        # Reformat time_keys to "Sept 2021", etc., if monthly
        if time_granularity == "month":
            pivot_df.columns = [
                datetime(int(year), int(month), 1).strftime("%b %Y")
                for (year, month) in pivot_df.columns
            ]
        elif time_granularity == "year":
            pivot_df.columns = [str(year) for year in pivot_df.columns]

        # Reset the index to make the 'Property' column appear explicitly
        pivot_df.reset_index(inplace=True)

        # Optionally, rename the 'Property' column to match property_granularity specifics
        pivot_df.rename(
            columns={
                "Property": granularity_label.get(property_granularity, "Property")
            },
            inplace=True,
        )

        return pivot_df

    def is_numeric(self, value):
        """
        Determines if a given value is numeric.
        Returns True if value is an integer or float, or a string that can be converted to a float.
        """
        if isinstance(value, (int, float)):
            return True
        if isinstance(value, str):
            value = value.strip()
            if value.isdigit():
                return True
            try:
                float(value)
                return True
            except ValueError:
                return False
        return False

    async def write_excel(
        self,
        df: pd.DataFrame,
        report_title: str = "Bed Night Report",
        include_total_column: bool = False,
    ):
        """Writes a dataframe into an Excel stream with enhanced formatting based on stakeholder requests."""
        excel_stream = BytesIO()
        header_row = 3
        tb_teal_fill = PatternFill(
            start_color="0E9BAC", end_color="0E9BAC", fill_type="solid"
        )
        tb_white_font = "F2F0E7"

        with pd.ExcelWriter(excel_stream, engine="openpyxl") as writer:
            # Write the DataFrame starting from row 3 to leave space for title and subtitle
            df.to_excel(
                writer, index=False, startrow=header_row - 1, sheet_name="Sheet1"
            )
            numeric_columns = df.select_dtypes(include=["number"]).columns.tolist()

            sheet = writer.sheets["Sheet1"]

            # Define the number of columns in the DataFrame
            df_width = len(df.columns)
            num_columns = df_width + 1 if include_total_column else df_width

            # ----------------------------
            # 1. Add Title (Row 1)
            # ----------------------------
            title_font = Font(name=FONT_NAME, bold=True, size=18, color=tb_white_font)
            sheet.merge_cells(
                start_row=1,
                start_column=1,
                end_row=1,
                end_column=num_columns,
            )
            title_cell = sheet.cell(row=1, column=1, value=report_title)
            title_cell.font = title_font
            title_cell.fill = tb_teal_fill
            title_cell.alignment = Alignment(horizontal="center", vertical="center")

            # ----------------------------
            # 2. Add Subtitle (Row 2)
            # ----------------------------
            subtitle_font = Font(name=FONT_NAME, size=14, color=tb_white_font)
            current_datetime = datetime.now()
            subtitle_text = f"Bed Nights as of {current_datetime.strftime('%B')} {current_datetime.day}, {current_datetime.year}"
            sheet.merge_cells(
                start_row=2, start_column=1, end_row=2, end_column=num_columns
            )
            subtitle_cell = sheet.cell(row=2, column=1, value=subtitle_text)
            subtitle_cell.font = subtitle_font
            subtitle_cell.fill = tb_teal_fill
            subtitle_cell.alignment = Alignment(horizontal="center", vertical="center")

            # ----------------------------
            # 3. Format Headers (Row 4)
            # ----------------------------

            header_font = Font(name=FONT_NAME, bold=True, size=11)
            if include_total_column:
                max_col = num_columns  # Include an extra column for 'TOTAL'
            else:
                max_col = num_columns  # Only include existing columns

            for col_num in range(1, max_col + 1):
                if col_num <= num_columns:
                    # Zero-based indexing in pandas
                    if col_num == num_columns and include_total_column:
                        header_value = "TOTAL"
                    else:
                        header_value = df.columns[col_num - 1]
                elif col_num == (num_columns - 1) and include_total_column:
                    # Last column when include_total_column is True
                    header_value = "TOTAL"

                cell = sheet.cell(row=header_row, column=col_num)

                if self.is_numeric(header_value):
                    # Convert to appropriate numeric type
                    if isinstance(header_value, float) or (
                        isinstance(header_value, str) and "." in header_value
                    ):
                        numeric_header = float(header_value)
                        cell.value = numeric_header
                        cell.number_format = "0.00"
                    else:
                        numeric_header = int(float(header_value))
                        cell.value = numeric_header
                        cell.number_format = "0"

                    # Set alignment for numeric headers
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                else:
                    # For non-numeric headers, set as uppercase string
                    cell.value = str(header_value).upper()
                    cell.number_format = "General"

                    # Set alignment based on header content
                    if cell.value == "PROPERTY" and include_total_column:
                        cell.alignment = Alignment(horizontal="left", vertical="center")
                    else:
                        cell.alignment = Alignment(
                            horizontal="center", vertical="center"
                        )

                # Apply font styling
                cell.font = header_font

            # ----------------------------
            # 4. Add TOTAL Column
            # ----------------------------
            total_col = num_columns
            if include_total_column:
                # total_col = num_columns + 1
                sheet.cell(row=header_row, column=total_col, value="TOTAL").font = (
                    header_font
                )
                sheet.cell(row=header_row, column=total_col).alignment = Alignment(
                    horizontal="center", vertical="center"
                )

                # Populate TOTAL column with formulas
                for row in range(header_row + 1, header_row + 1 + len(df)):
                    # Assuming numerical data starts from column 2 to (total_col -1)
                    sum_range = f"${get_column_letter(2)}${row}:${get_column_letter(num_columns-1)}${row}"
                    total_cell = sheet.cell(row=row, column=total_col)
                    total_cell.value = f"=SUM({sum_range})"
                    total_cell.font = Font(name=FONT_NAME, bold=True, size=11)
                    total_cell.alignment = Alignment(
                        horizontal="center", vertical="center"
                    )

            # ----------------------------
            # 5. Add TOTAL Row
            # ----------------------------
            total_row = header_row + 1 + len(df)
            sheet.cell(row=total_row, column=1, value="TOTAL").font = Font(
                name=FONT_NAME, bold=True, size=11
            )
            sheet.cell(row=total_row, column=1).alignment = Alignment(
                horizontal="right", vertical="center"
            )
            for col_num in range(2, total_col + 1):
                if include_total_column and col_num == total_col:
                    # Grand total for TOTAL column
                    sum_range = f"${get_column_letter(col_num)}${header_row + 1}:${get_column_letter(col_num)}${total_row - 1}"
                    total_cell = sheet.cell(row=total_row, column=col_num)
                    total_cell.value = f"=SUM({sum_range})"
                    total_cell.font = Font(name=FONT_NAME, bold=True, size=11)
                else:
                    column_name = df.columns[col_num - 1]
                    if column_name in numeric_columns:
                        sum_range = f"${get_column_letter(col_num)}${header_row + 1}:${get_column_letter(col_num)}${total_row - 1}"
                        total_cell = sheet.cell(row=total_row, column=col_num)
                        total_cell.value = f"=SUM({sum_range})"
                        total_cell.font = Font(name=FONT_NAME, bold=True, size=11)
                    else:
                        total_cell = sheet.cell(row=total_row, column=col_num)
                        total_cell.value = ""
                        total_cell.font = Font(name=FONT_NAME, bold=False, size=11)
                # Set alignment for all cells in TOTAL row
                total_cell.alignment = Alignment(horizontal="center", vertical="center")

            # ----------------------------
            # 6. Apply Gridlines (Borders) to the Table
            # ----------------------------
            thin_border = Border(
                left=Side(style="thin", color="000000"),
                right=Side(style="thin", color="000000"),
                top=Side(style="thin", color="000000"),
                bottom=Side(style="thin", color="000000"),
            )

            for row in range(header_row, total_row + 2):
                for col in range(1, num_columns + 1):
                    cell = sheet.cell(row=row, column=col)
                    cell.border = thin_border
                    if row == header_row:
                        # Header cells already styled
                        pass
                    elif row == total_row:
                        # Total cells already styled
                        pass
                    elif row >= header_row + 1 and row < total_row and col != total_col:
                        # Table data cells
                        cell.font = Font(name=FONT_NAME, size=11)
                        column_name = df.columns[col - 1]
                        if column_name in numeric_columns:
                            cell.alignment = Alignment(
                                horizontal="center", vertical="center"
                            )
                            cell.number_format = "General"
                    elif include_total_column and col == total_col:
                        # TOTAL column cells
                        cell.font = Font(name=FONT_NAME, bold=True, size=11)

            # ----------------------------
            # 7. Set Column Widths
            # ----------------------------
            for col in sheet.columns:
                max_length = 0
                column = None  # Initialize the column variable

                # Find the first non-merged cell in the column to get the column letter
                for cell in col:
                    if not isinstance(cell, MergedCell):
                        column = cell.column_letter
                        break

                if column is None:
                    # All cells in the column are merged; skip adjusting this column
                    continue

                for cell in col:
                    if isinstance(cell, MergedCell):
                        # Skip MergedCell objects
                        continue
                    try:
                        cell_value = str(cell.value) if cell.value is not None else ""
                        cell_length = len(cell_value)
                        if cell_length > max_length:
                            max_length = cell_length
                    except:
                        pass
                adjusted_width = max_length + 2  # Adding extra space for readability
                sheet.column_dimensions[column].width = adjusted_width
            # ----------------------------
            # 8. Add Final Footer (Optional)
            # ----------------------------
            footer_row = total_row + 1
            sheet.cell(row=footer_row, column=1, value="Travel Beyond Confidential")
            footer_font = Font(
                name=FONT_NAME,
                bold=True,
                size=14,
                color=tb_white_font,
            )
            footer_alignment = Alignment(horizontal="center", vertical="center")

            sheet.merge_cells(
                start_row=footer_row,
                start_column=1,
                end_row=footer_row,
                end_column=total_col,
            )
            footer_cell = sheet.cell(row=footer_row, column=1)
            footer_cell.font = footer_font
            footer_cell.fill = tb_teal_fill
            footer_cell.alignment = footer_alignment
            footer_cell.border = thin_border

            for col_num in range(1, total_col + 1):
                cell = sheet.cell(row=footer_row, column=col_num)
                cell.border = thin_border

        excel_stream.seek(0)  # Rewind the buffer to the beginning after writing
        return excel_stream

    def aggregate_custom_report(
        self,
        accommodation_logs: Sequence[AccommodationLogSummary],
        calculation_type: str,
        property_granularity: str,
        time_granularity: str,
    ):
        """Performs calculations for custom report results."""
        # Define key functions for grouping by time
        if time_granularity == "month":
            key_func = lambda x: (x.date_in.year, x.date_in.month)
        elif time_granularity == "year":
            key_func = lambda x: x.date_in.year

        # Group logs by time
        logs_grouped_by_time = defaultdict(list)
        for key, group in groupby(
            sorted(accommodation_logs, key=key_func), key=key_func
        ):
            logs_grouped_by_time[key].extend(group)

        # Now aggregate within each time group based on property granularity
        final_aggregation = {}
        for time_key, logs in logs_grouped_by_time.items():
            group_key_func = attrgetter(
                property_granularity
            )  # e.g., 'property_name', 'property_portfolio', 'country_name'
            property_group = defaultdict(list)
            for prop_key, prop_group in groupby(
                sorted(logs, key=group_key_func), key=group_key_func
            ):
                property_group[prop_key].extend(prop_group)

            # Aggregate counts or bed nights within each property group
            for prop_key, prop_logs in property_group.items():
                if calculation_type == "bed_nights":
                    total = sum(log.bed_nights for log in prop_logs)
                elif calculation_type == "num_bookings":
                    total = len(prop_logs)

                # Store aggregated data
                final_aggregation[(time_key, prop_key)] = total

        return final_aggregation

    # AccommodationLog
    async def get_all_accommodation_logs(self) -> Sequence[AccommodationLogSummary]:
        """Gets all AccommodationLogSummary models."""
        return await self._repo.get_all_accommodation_logs()

    async def get_accommodation_logs_by_filters(
        self, filters: dict
    ) -> Sequence[AccommodationLogSummary]:
        """Gets all AccommodationLogSummary models based on a filter."""
        return await self._repo.get_accommodation_logs_by_filter(filters)

    async def get_related_records_summary(
        self, identifier: UUID, identifier_type: str
    ) -> Sequence[AccommodationLogSummary] | dict:
        """Checks the impact of a modification on related records."""
        filters = {identifier_type: identifier}
        accommodation_logs = await self.get_accommodation_logs_by_filters(filters)

        if not accommodation_logs:
            return {"can_modify": True, "affected_logs": []}

        accommodation_logs_summary = [log.to_json() for log in accommodation_logs]
        return {"can_modify": False, "affected_logs": accommodation_logs_summary}

    async def get_overlaps(self, start_date: date, end_date: date) -> list[str]:
        """Gets records where clients will be overlapping."""
        overlap_data = await self._repo.get_overlaps(start_date, end_date)
        overlaps = [overlap.to_json() for overlap in overlap_data]
        return overlaps

    # Property
    async def get_all_properties(self) -> Sequence[PropertySummary]:
        """Gets all PropertySummary models."""
        return await self._repo.get_all_properties()

    # Property
    async def get_all_property_details(self) -> Sequence[PropertyDetailSummary]:
        """Gets all PropertyDetailSummary models."""
        return await self._repo.get_property_details()

    async def get_property_details_by_id(
        self, property_id: UUID
    ) -> Optional[PropertyDetailSummary]:
        """Gets PropertyDetailSummary model by ID."""
        res = await self._repo.get_property_details(
            entered_only=False, by_id=property_id
        )
        if res:
            return res[0]
        else:
            return None

    # Country
    async def get_all_countries(self) -> Sequence[CountrySummary]:
        """Gets all CountrySummary models."""
        return await self._repo.get_all_countries()

    async def get_country_details_by_id(
        self, country_id: UUID
    ) -> Sequence[CountrySummary]:
        """Gets CountrySummary model by ID."""
        return await self._repo.get_country_by_id(country_id)

    # Agency
    async def get_all_agencies(self) -> Sequence[AgencySummary]:
        """Gets all AgencySummary models."""
        return await self._repo.get_all_agencies()

    # BookingChannel
    async def get_all_booking_channels(self) -> Sequence[BookingChannelSummary]:
        """Gets all BookingChannelSummary models."""
        return await self._repo.get_all_booking_channels()

    # Portfolio
    async def get_all_portfolios(self) -> Sequence[PortfolioSummary]:
        """Gets all PortfolioSummary models."""
        return await self._repo.get_all_portfolios()

    # Trip
    async def get_all_trips(self) -> Sequence[TripSummary]:
        """Gets all TripSummary models."""
        return await self._repo.get_all_trips()

    async def get_trip_summary_by_id(self, trip_id: UUID) -> TripSummary:
        """Gets a TripSummary model by ID."""
        return await self._repo.get_trip_summary_by_id(trip_id)
